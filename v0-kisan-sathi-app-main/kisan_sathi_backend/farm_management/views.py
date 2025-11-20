from django.shortcuts import render
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import datetime, timedelta
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.template.loader import get_template
from django.contrib.auth.decorators import login_required
from rest_framework_simplejwt.authentication import JWTAuthentication
from functools import wraps
import json
from decimal import Decimal
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

from .models import (
    ExpenseCategory, Expense, Crop, Income, InventoryCategory, 
    InventoryItem, CropPlan, LivestockType, Livestock, 
    VaccinationRecord, Loan, EMIPayment
)
from .serializers import (
    ExpenseCategorySerializer, ExpenseSerializer, CropSerializer, 
    IncomeSerializer, InventoryCategorySerializer, InventoryItemSerializer,
    CropPlanSerializer, LivestockTypeSerializer, LivestockSerializer,
    VaccinationRecordSerializer, LoanSerializer, EMIPaymentSerializer,
    MonthlyProfitSerializer, DashboardStatsSerializer, ExpenseByCategorySerializer
)
from .utils import format_currency, register_unicode_fonts, format_excel_currency

# Custom decorator for dual authentication (Session + JWT)
def dual_auth_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        # Try JWT authentication first
        jwt_auth = JWTAuthentication()
        try:
            user_auth_tuple = jwt_auth.authenticate(request)
            if user_auth_tuple is not None:
                request.user, request.auth = user_auth_tuple
                return view_func(request, *args, **kwargs)
        except:
            pass
        
        # Fall back to session authentication
        if request.user.is_authenticated:
            return view_func(request, *args, **kwargs)
        
        # Not authenticated
        return HttpResponse('Unauthorized', status=401)
    return wrapper

class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    queryset = ExpenseCategory.objects.all()
    serializer_class = ExpenseCategorySerializer
    permission_classes = []  # Public access for reference data

class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Expense.objects.filter(farmer=self.request.user)
        
        # Filter by category
        category = self.request.query_params.get('category', None)
        if category:
            queryset = queryset.filter(category_id=category)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        # Search by notes
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(Q(notes__icontains=search) | Q(category__name__icontains=search))
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(farmer=self.request.user)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response({
            'success': True,
            'message': 'Expense added successfully',
            'data': serializer.data
        }, status=status.HTTP_201_CREATED, headers=headers)
    
    def destroy(self, request, *args, **kwargs):
        """Delete expense with proper JSON response"""
        try:
            instance = self.get_object()
            # Verify ownership
            if instance.farmer != request.user:
                return Response({
                    'success': False,
                    'message': 'You do not have permission to delete this expense'
                }, status=status.HTTP_403_FORBIDDEN)
            
            self.perform_destroy(instance)
            return Response({
                'success': True,
                'message': 'Expense deleted successfully'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'success': False,
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        queryset = self.get_queryset()
        total = queryset.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        count = queryset.count()
        return Response({
            'total_expenses': total,
            'expense_count': count
        })

class CropViewSet(viewsets.ModelViewSet):
    queryset = Crop.objects.all()
    serializer_class = CropSerializer
    permission_classes = []  # Public access for reference data

class IncomeViewSet(viewsets.ModelViewSet):
    serializer_class = IncomeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Income.objects.filter(farmer=self.request.user)
        
        # Filter by crop
        crop = self.request.query_params.get('crop', None)
        if crop:
            queryset = queryset.filter(crop_id=crop)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        if start_date:
            queryset = queryset.filter(sale_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(sale_date__lte=end_date)
        
        # Search by buyer name or crop
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(Q(buyer_name__icontains=search) | Q(crop__name__icontains=search))
        
        # Filter by payment status
        payment_status = self.request.query_params.get('payment_status', None)
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(farmer=self.request.user)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response({
            'success': True,
            'message': 'Income recorded successfully',
            'data': serializer.data
        }, status=status.HTTP_201_CREATED, headers=headers)
    
    def destroy(self, request, *args, **kwargs):
        """Delete income with proper JSON response"""
        try:
            instance = self.get_object()
            # Verify ownership
            if instance.farmer != request.user:
                return Response({
                    'success': False,
                    'message': 'You do not have permission to delete this income'
                }, status=status.HTTP_403_FORBIDDEN)
            
            self.perform_destroy(instance)
            return Response({
                'success': True,
                'message': 'Income deleted successfully'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'success': False,
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        queryset = self.get_queryset()
        total = queryset.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        count = queryset.count()
        return Response({
            'total_income': total,
            'income_count': count
        })

class InventoryCategoryViewSet(viewsets.ModelViewSet):
    queryset = InventoryCategory.objects.all()
    serializer_class = InventoryCategorySerializer
    permission_classes = []  # Public access for reference data

class InventoryItemViewSet(viewsets.ModelViewSet):
    serializer_class = InventoryItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = InventoryItem.objects.filter(farmer=self.request.user)
        
        # Filter by category
        category = self.request.query_params.get('category', None)
        if category:
            queryset = queryset.filter(category_id=category)
        
        # Search by name
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(brand__icontains=search))
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(farmer=self.request.user)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response({
            'success': True,
            'message': 'Inventory item added successfully',
            'data': serializer.data
        }, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        low_stock_items = self.get_queryset().filter(
            current_stock__lte=F('minimum_stock')
        )
        serializer = self.get_serializer(low_stock_items, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def use_stock(self, request, pk=None):
        item = self.get_object()
        quantity = request.data.get('quantity', 0)
        
        try:
            quantity = Decimal(str(quantity))
            if quantity <= 0:
                return Response({
                    'success': False,
                    'message': 'Quantity must be greater than 0'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if item.current_stock < quantity:
                return Response({
                    'success': False,
                    'message': 'Insufficient stock'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            item.current_stock -= quantity
            item.save()
            
            serializer = self.get_serializer(item)
            return Response({
                'success': True,
                'message': f'Used {quantity} {item.unit} of {item.name}',
                'data': serializer.data
            })
        except (ValueError, TypeError):
            return Response({
                'success': False,
                'message': 'Invalid quantity'
            }, status=status.HTTP_400_BAD_REQUEST)

class CropPlanViewSet(viewsets.ModelViewSet):
    serializer_class = CropPlanSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return CropPlan.objects.filter(farmer=self.request.user)

    def perform_create(self, serializer):
        serializer.save(farmer=self.request.user)

class LivestockTypeViewSet(viewsets.ModelViewSet):
    queryset = LivestockType.objects.all()
    serializer_class = LivestockTypeSerializer
    permission_classes = []  # Public access for reference data

class LivestockViewSet(viewsets.ModelViewSet):
    serializer_class = LivestockSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Livestock.objects.filter(farmer=self.request.user)

    def perform_create(self, serializer):
        serializer.save(farmer=self.request.user)

class VaccinationRecordViewSet(viewsets.ModelViewSet):
    serializer_class = VaccinationRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return VaccinationRecord.objects.filter(livestock__farmer=self.request.user)

    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        upcoming_date = timezone.now().date() + timedelta(days=30)
        upcoming_vaccinations = self.get_queryset().filter(
            next_due_date__lte=upcoming_date,
            next_due_date__gte=timezone.now().date()
        )
        serializer = self.get_serializer(upcoming_vaccinations, many=True)
        return Response(serializer.data)

class LoanViewSet(viewsets.ModelViewSet):
    serializer_class = LoanSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Loan.objects.filter(farmer=self.request.user)

    def perform_create(self, serializer):
        serializer.save(farmer=self.request.user)

class EMIPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = EMIPaymentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return EMIPayment.objects.filter(loan__farmer=self.request.user)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    farmer = request.user
    current_year = timezone.now().year
    
    # Calculate total income and expenses for current year
    total_income = Income.objects.filter(
        farmer=farmer, 
        sale_date__year=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    total_expenses = Expense.objects.filter(
        farmer=farmer, 
        date__year=current_year
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    net_profit = total_income - total_expenses
    
    # Count active loans
    active_loans = Loan.objects.filter(farmer=farmer, status='active').count()
    
    # Count low stock items
    low_stock_items = InventoryItem.objects.filter(
        farmer=farmer,
        current_stock__lte=F('minimum_stock')
    ).count()
    
    # Count upcoming vaccinations (next 30 days)
    upcoming_date = timezone.now().date() + timedelta(days=30)
    upcoming_vaccinations = VaccinationRecord.objects.filter(
        livestock__farmer=farmer,
        next_due_date__lte=upcoming_date,
        next_due_date__gte=timezone.now().date()
    ).count()
    
    # Count active crop plans
    active_crop_plans = CropPlan.objects.filter(
        farmer=farmer, 
        status__in=['planned', 'planted', 'growing']
    ).count()
    
    stats = {
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'active_loans': active_loans,
        'low_stock_items': low_stock_items,
        'upcoming_vaccinations': upcoming_vaccinations,
        'active_crop_plans': active_crop_plans
    }
    
    serializer = DashboardStatsSerializer(stats)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def monthly_profit_chart(request):
    farmer = request.user
    year = request.GET.get('year', timezone.now().year)
    
    monthly_data = []
    
    for month in range(1, 13):
        # Calculate monthly income
        monthly_income = Income.objects.filter(
            farmer=farmer,
            sale_date__year=year,
            sale_date__month=month
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        # Calculate monthly expenses
        monthly_expense = Expense.objects.filter(
            farmer=farmer,
            date__year=year,
            date__month=month
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        monthly_profit = monthly_income - monthly_expense
        
        monthly_data.append({
            'month': datetime(int(year), month, 1).strftime('%B'),
            'total_income': monthly_income,
            'total_expense': monthly_expense,
            'profit': monthly_profit
        })
    
    serializer = MonthlyProfitSerializer(monthly_data, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def expense_by_category(request):
    farmer = request.user
    year = request.GET.get('year', timezone.now().year)
    
    # Get expenses by category for the year
    category_expenses = Expense.objects.filter(
        farmer=farmer,
        date__year=year
    ).values('category__name').annotate(
        total_amount=Sum('amount')
    ).order_by('-total_amount')
    
    # Calculate total expenses for percentage calculation
    total_expenses = sum(item['total_amount'] for item in category_expenses)
    
    # Add percentage to each category
    for item in category_expenses:
        if total_expenses > 0:
            item['percentage'] = (item['total_amount'] / total_expenses) * 100
        else:
            item['percentage'] = 0
        item['category'] = item['category__name']
    
    serializer = ExpenseByCategorySerializer(category_expenses, many=True)
    return Response(serializer.data)

# Template Views for Frontend
def dashboard_view(request):
    return render(request, 'farm_management/dashboard.html')

def expenses_view(request):
    return render(request, 'farm_management/expenses.html')

def income_view(request):
    return render(request, 'farm_management/income.html')

def inventory_view(request):
    return render(request, 'farm_management/inventory.html')

def crop_planning_view(request):
    return render(request, 'farm_management/crop_planning.html')

def livestock_view(request):
    return render(request, 'farm_management/livestock.html')

def loans_view(request):
    return render(request, 'farm_management/loans.html')

def reports_view(request):
    return render(request, 'farm_management/reports.html')

# PDF Export
@dual_auth_required
def export_expenses_pdf(request):
    farmer = request.user
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Register Unicode fonts for rupee symbol support
    font_name = register_unicode_fonts()
    
    # Query expenses
    expenses = Expense.objects.filter(farmer=farmer)
    if start_date:
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        expenses = expenses.filter(date__lte=end_date)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title with Unicode font
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=font_name
    )
    elements.append(Paragraph('Expense Report', title_style))
    elements.append(Spacer(1, 12))
    
    # Farmer info with Unicode font
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontName=font_name
    )
    elements.append(Paragraph(f'<b>Farmer:</b> {farmer.get_full_name() or farmer.username}', info_style))
    elements.append(Paragraph(f'<b>Report Date:</b> {timezone.now().strftime("%d %B %Y")}', info_style))
    if start_date and end_date:
        elements.append(Paragraph(f'<b>Period:</b> {start_date} to {end_date}', info_style))
    elements.append(Spacer(1, 20))
    
    # Table data with formatted currency
    data = [['Date', 'Category', 'Amount', 'Notes']]
    total = Decimal('0')
    
    for expense in expenses:
        data.append([
            expense.date.strftime('%d-%m-%Y'),
            expense.category.name,
            format_currency(expense.amount),
            expense.notes[:50] if expense.notes else '-'
        ])
        total += expense.amount
    
    # Add total row
    data.append(['', '', f'Total: {format_currency(total)}', ''])
    
    # Create table with Unicode font
    table = Table(data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME', (0, -1), (-1, -1), font_name),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="expenses_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response

@dual_auth_required
def export_income_pdf(request):
    farmer = request.user
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Register Unicode fonts for rupee symbol support
    font_name = register_unicode_fonts()
    
    # Query income
    income_records = Income.objects.filter(farmer=farmer)
    if start_date:
        income_records = income_records.filter(sale_date__gte=start_date)
    if end_date:
        income_records = income_records.filter(sale_date__lte=end_date)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title with Unicode font
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#27ae60'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=font_name
    )
    elements.append(Paragraph('Income Report', title_style))
    elements.append(Spacer(1, 12))
    
    # Farmer info with Unicode font
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontName=font_name
    )
    elements.append(Paragraph(f'<b>Farmer:</b> {farmer.get_full_name() or farmer.username}', info_style))
    elements.append(Paragraph(f'<b>Report Date:</b> {timezone.now().strftime("%d %B %Y")}', info_style))
    if start_date and end_date:
        elements.append(Paragraph(f'<b>Period:</b> {start_date} to {end_date}', info_style))
    elements.append(Spacer(1, 20))
    
    # Table data with formatted currency
    data = [['Date', 'Crop', 'Quantity', 'Rate', 'Total', 'Buyer']]
    total = Decimal('0')
    
    for income in income_records:
        data.append([
            income.sale_date.strftime('%d-%m-%Y'),
            income.crop.name,
            f'{income.quantity} {income.unit}',
            format_currency(income.rate_per_unit),
            format_currency(income.total_amount),
            income.buyer_name[:20]
        ])
        total += income.total_amount
    
    # Add total row
    data.append(['', '', '', '', f'Total: {format_currency(total)}', ''])
    
    # Create table with Unicode font
    table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1.2*inch, 1.2*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME', (0, -1), (-1, -1), font_name),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="income_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response

# Excel Export
@dual_auth_required
def export_expenses_excel(request):
    farmer = request.user
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Query expenses
    expenses = Expense.objects.filter(farmer=farmer)
    if start_date:
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        expenses = expenses.filter(date__lte=end_date)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'
    
    # Header styling
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # Headers
    headers = ['Date', 'Category', 'Amount', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data with currency formatting
    total = Decimal('0')
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=expense.date.strftime('%d-%m-%Y'))
        ws.cell(row=row, column=2, value=expense.category.name)
        
        # Format currency cell
        amount_cell = ws.cell(row=row, column=3, value=float(expense.amount))
        amount_cell.number_format = '₹#,##0.00'
        
        ws.cell(row=row, column=4, value=expense.notes)
        total += expense.amount
    
    # Total row with currency formatting
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=2, value='TOTAL').font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=3, value=float(total))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '₹#,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="expenses_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response

@dual_auth_required
def export_income_excel(request):
    farmer = request.user
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Query income
    income_records = Income.objects.filter(farmer=farmer)
    if start_date:
        income_records = income_records.filter(sale_date__gte=start_date)
    if end_date:
        income_records = income_records.filter(sale_date__lte=end_date)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Income'
    
    # Header styling
    header_fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # Headers
    headers = ['Date', 'Crop', 'Quantity', 'Unit', 'Rate per Unit', 'Total Amount', 'Buyer', 'Payment Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data with currency formatting
    total = Decimal('0')
    for row, income in enumerate(income_records, 2):
        ws.cell(row=row, column=1, value=income.sale_date.strftime('%d-%m-%Y'))
        ws.cell(row=row, column=2, value=income.crop.name)
        ws.cell(row=row, column=3, value=float(income.quantity))
        ws.cell(row=row, column=4, value=income.unit)
        
        # Format currency cells
        rate_cell = ws.cell(row=row, column=5, value=float(income.rate_per_unit))
        rate_cell.number_format = '₹#,##0.00'
        
        amount_cell = ws.cell(row=row, column=6, value=float(income.total_amount))
        amount_cell.number_format = '₹#,##0.00'
        
        ws.cell(row=row, column=7, value=income.buyer_name)
        ws.cell(row=row, column=8, value=income.payment_status)
        total += income.total_amount
    
    # Total row with currency formatting
    total_row = len(income_records) + 2
    ws.cell(row=total_row, column=5, value='TOTAL').font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=6, value=float(total))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '₹#,##0.00'
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="income_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response

@dual_auth_required
def export_analytics_pdf(request):
    farmer = request.user
    year = request.GET.get('year', timezone.now().year)
    
    # Register Unicode fonts for rupee symbol support
    font_name = register_unicode_fonts()
    
    # Get analytics data
    total_income = Income.objects.filter(
        farmer=farmer, 
        sale_date__year=year
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    total_expenses = Expense.objects.filter(
        farmer=farmer, 
        date__year=year
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    net_profit = total_income - total_expenses
    
    # Category-wise expenses
    category_expenses = Expense.objects.filter(
        farmer=farmer,
        date__year=year
    ).values('category__name').annotate(
        total_amount=Sum('amount')
    ).order_by('-total_amount')
    
    # Category-wise income (by crop)
    category_income = Income.objects.filter(
        farmer=farmer,
        sale_date__year=year
    ).values('crop__name').annotate(
        total_amount=Sum('total_amount')
    ).order_by('-total_amount')
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title with Unicode font
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#8e44ad'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=font_name
    )
    elements.append(Paragraph(f'Farm Analytics Report - {year}', title_style))
    elements.append(Spacer(1, 12))
    
    # Farmer info with Unicode font
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontName=font_name
    )
    elements.append(Paragraph(f'<b>Farmer:</b> {farmer.get_full_name() or farmer.username}', info_style))
    elements.append(Paragraph(f'<b>Report Date:</b> {timezone.now().strftime("%d %B %Y")}', info_style))
    elements.append(Spacer(1, 20))
    
    # Summary table with formatted currency
    summary_data = [
        ['Metric', 'Amount'],
        ['Total Income', format_currency(total_income)],
        ['Total Expenses', format_currency(total_expenses)],
        ['Net Profit/Loss', format_currency(net_profit)]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8e44ad')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f39c12') if net_profit >= 0 else colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, -1), (-1, -1), font_name),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Category-wise expenses with Unicode font
    heading_style = ParagraphStyle(
        'HeadingStyle',
        parent=styles['Heading2'],
        fontName=font_name
    )
    elements.append(Paragraph('<b>Expense Breakdown by Category</b>', heading_style))
    elements.append(Spacer(1, 12))
    
    category_data = [['Category', 'Amount', 'Percentage']]
    for cat in category_expenses:
        percentage = (cat['total_amount'] / total_expenses * 100) if total_expenses > 0 else 0
        category_data.append([
            cat['category__name'],
            format_currency(cat['total_amount']),
            f'{percentage:.1f}%'
        ])
    
    category_table = Table(category_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
    category_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(category_table)
    elements.append(Spacer(1, 30))
    
    # Income Breakdown by Category with Unicode font
    elements.append(Paragraph('<b>Income Breakdown by Category</b>', heading_style))
    elements.append(Spacer(1, 12))
    
    income_data = [['Category (Crop)', 'Amount', 'Percentage']]
    for crop in category_income:
        percentage = (crop['total_amount'] / total_income * 100) if total_income > 0 else 0
        income_data.append([
            crop['crop__name'],
            format_currency(crop['total_amount']),
            f'{percentage:.1f}%'
        ])
    
    income_table = Table(income_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
    income_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(income_table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="analytics_{year}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response


# Income Breakdown View
from django.contrib.auth.decorators import login_required

@login_required
def income_breakdown_view(request):
    """Income breakdown by category view"""
    return render(request, 'farm_management/income_breakdown.html')

@login_required
def income_expense_combined_view(request):
    """Combined income and expense management page"""
    return render(request, 'farm_management/income_expense_combined.html')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def income_by_crop(request):
    """Get income breakdown by crop"""
    farmer = request.user
    year = request.GET.get('year', timezone.now().year)
    
    # Get income by crop for the year
    crop_income = Income.objects.filter(
        farmer=farmer,
        sale_date__year=year
    ).values('crop__name').annotate(
        total_amount=Sum('total_amount')
    ).order_by('-total_amount')
    
    # Calculate total income for percentage
    total_income = sum(item['total_amount'] for item in crop_income)
    
    # Add percentage to each crop
    for item in crop_income:
        if total_income > 0:
            item['percentage'] = (item['total_amount'] / total_income) * 100
        else:
            item['percentage'] = 0
        item['crop'] = item['crop__name']
    
    return Response(list(crop_income))
